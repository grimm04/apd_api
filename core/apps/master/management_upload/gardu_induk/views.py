import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik,fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete

class GarduIndukView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=4)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_gardu_induk',
        summary="Management Upload SLD - Upload Temp Excel Gardu Induk",
        description="Management Upload SLD - Upload Temp Excel Gardu Induk",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_GI']),
                    kode_lokasi=fillnan(value=row['KODE_GI']),
                    id_parent_lokasi=fillnan(value=row['ID_PARENT_LOKASI']),
                    tree_jaringan=1,
                    id_ref_jenis_lokasi=4,
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    id_up2b=fillnan(value=row['ID_UP2B']),
                    jenis_gi=fillnan(value=row['JENIS_GI']),
                    fungsi_scada=fillnan(value=row['FUNGSI_SCADA']),
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Gardu Induk",
        description="Management Upload SLD - Download Template Excel Gardu Induk",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='gardu_induk')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_gardu_induk.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            data_upload.set_column('A:A', 20)
            data_upload.set_column('B:B', 25)
            data_upload.set_column('C:C', 50)
            data_upload.set_column('D:D', 20)
            data_upload.set_column('E:E', 15)
            data_upload.set_column('F:F', 15)
            data_upload.set_column('G:G', 15)
            data_upload.set_column('H:H', 15)
            data_upload.set_column('I:I', 15)
            data_upload.set_column('J:J', 15)
            data_upload.set_column('K:K', 15)
            data_upload.set_column('L:K', 15)
            data_upload.write('A1', "ID_GI", format)
            data_upload.write('B1', "KODE_GI", format)
            data_upload.write('C1', "NAMA_GI", format)
            data_upload.write('D1', "ALAMAT", format)
            data_upload.write('E1', "ID_PARENT_LOKASI", format)
            data_upload.write('F1', "LATITUDE", format)
            data_upload.write('G1', "LONGITUDE", format)
            data_upload.write('H1', "ID_UP2B", format)
            data_upload.write('I1', "JENIS_GI", format)
            data_upload.write('J1', "FUNGSI_SCADA", format)
            data_upload.write('K1', "STATUS", format)
            data_upload.write('L1', "EVENT_UPLOAD", format)
            sample_upload.set_tab_color('red')
            sample_upload.set_column('A:A', 20)
            sample_upload.set_column('B:B', 25)
            sample_upload.set_column('C:C', 50)
            sample_upload.set_column('D:D', 20)
            sample_upload.set_column('E:E', 15)
            sample_upload.set_column('F:F', 15)
            sample_upload.set_column('G:G', 15)
            sample_upload.set_column('H:H', 15)
            sample_upload.set_column('I:I', 15)
            sample_upload.set_column('J:J', 15)
            sample_upload.set_column('K:K', 15)
            sample_upload.set_column('L:K', 15)

            sample_upload.write('A1', "ID_GI", format)
            sample_upload.write('B1', "KODE_GI", format)
            sample_upload.write('C1', "NAMA_GI", format)
            sample_upload.write('D1', "ALAMAT", format)
            sample_upload.write('E1', "ID_PARENT_LOKASI", format)
            sample_upload.write('F1', "LATITUDE", format)
            sample_upload.write('G1', "LONGITUDE", format)
            sample_upload.write('H1', "ID_UP2B", format)
            sample_upload.write('I1', "JENIS_GI", format)
            sample_upload.write('J1', "FUNGSI_SCADA", format)
            sample_upload.write('K1', "STATUS", format)
            sample_upload.write('L1', "EVENT_UPLOAD", format)
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_gi'))
                sample_upload.write(index, 1, data.get('kode_lokasi'))
                sample_upload.write(index, 2, data.get('nama_gi'))
                sample_upload.write(index, 3, data.get('alamat'))
                sample_upload.write(index, 4, data.get('id_parent_lokasi'))
                sample_upload.write(index, 5, data.get('latitude'))
                sample_upload.write(index, 6, data.get('longitude'))
                sample_upload.write(index, 7, data.get('id_up2b'))
                sample_upload.write(index, 8, data.get('jenis_gi'))
                sample_upload.write(index, 9, data.get('fungsi_scada'))
                sample_upload.write(index, 10, data.get('status'))
                sample_upload.write(index, 11, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_GI UNTUK EVENT UPLOAD "INSERT"', merge_format)
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            _, message = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=2)
            if _: 
                ref_jenis_lokasi = workbook.add_worksheet('REF PEMBANGKIT')
                ref_jenis_lokasi.set_tab_color('red')
                ref_jenis_lokasi.set_column('A:A', 15)
                ref_jenis_lokasi.set_column('B:B', 25)
                ref_jenis_lokasi.write('A1', "ID_PEMBANGKIT", format)
                ref_jenis_lokasi.write('B1', "NAMA_PEMBANGKIT", format)
                index = 1
                for message in message:
                    ref_jenis_lokasi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_jenis_lokasi.write(index, 1, message['nama_lokasi'])
                    index += 1 
                ref_jenis_lokasi.merge_range('D2:N2', 'ID_PEMBANGKIT DIGUNAKAN UNTUK MENGISI ID_PARENT_LOKASI KETIKA UPLOAD', merge_format)


            _, message = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=25)
            if  _: 
                up2b = workbook.add_worksheet('REF UP2B')
                up2b.set_tab_color('red')
                up2b.set_column('A:A', 15)
                up2b.set_column('B:B', 25)
                up2b.write('A1', "ID_UP2B", format)
                up2b.write('B1', "NAMA_UP2B", format)
                index = 1
                for message in message:
                    up2b.write(index, 0, str(message['id_ref_lokasi']))
                    up2b.write(index, 1, message['nama_lokasi'])
                    index += 1 
                up2b.merge_range('D2:N2', 'ID_UP2B DIGUNAKAN UNTUK MENGISI ID_UP2B KETIKA UPLOAD', merge_format)

            #jenis gi
            jg = workbook.add_worksheet('REF JENIS GI')  
            jg.set_tab_color('red')
            jg.set_column('A:A', 25) 
            jg.write('A1', "NAMA_JENIS_GI", format)
            index = 1
            jenis_gi = ['GI','GIS']
            for message in jenis_gi:
                jg.write(index, 0, str(message)) 
                index += 1
            
            jg.merge_range('D2:N2', 'NAMA_JENIS_GI DIGUNAKAN UNTUK MENGISI JENIS_GI KETIKA UPLOAD', merge_format)

            
            #fungsi scada
            fs = workbook.add_worksheet('REF FUNGSI SCADA')  
            fs.set_tab_color('red')
            fs.set_column('A:A', 25) 
            fs.write('A1', "NAMA_FUNGSI_SCADA", format)
            index = 1
            fungsi_scada = ['GI','GH','KP','FI','EVM','MASTER']
            for message in fungsi_scada:
                fs.write(index, 0, str(message)) 
                index += 1
            fs.merge_range('D2:N2', 'NAMA_FUNGSI_SCADA DIGUNAKAN UNTUK MENGISI FUNGSI_SCADA KETIKA UPLOAD', merge_format)

            
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Gardu Induk",
        description="Management Upload SLD - Download Data Excel Gardu Induk",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD MASTER GARDU INDUK.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA GARDU INDUK')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            ws_acc.set_column('A:A', 20)
            ws_acc.set_column('B:B', 25) 
            ws_acc.set_column('C:C', 50)
            ws_acc.set_column('D:D', 20)
            ws_acc.set_column('E:E', 25)
            ws_acc.set_column('F:F', 15)
            ws_acc.set_column('G:G', 15)
            ws_acc.set_column('H:H', 15)
            ws_acc.set_column('I:I', 15)
            ws_acc.set_column('J:J', 15)
            ws_acc.set_column('K:K', 15) 
            ws_acc.write('A1', "ID_GI", format)
            ws_acc.write('B1', "KODE_GI", format)
            ws_acc.write('C1', "NAMA_GI", format)
            ws_acc.write('D1', "ALAMAT", format)
            ws_acc.write('E1', "ID_PARENT_LOKASI", format)
            ws_acc.write('F1', "LATITUDE", format)
            ws_acc.write('G1', "LONGITUDE", format)
            ws_acc.write('H1', "ID_UP2B", format)
            ws_acc.write('I1', "JENIS_GI", format)
            ws_acc.write('J1', "FUNGSI_SCADA", format)
            ws_acc.write('K1', "STATUS", format) 
            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('up2b'):
                    up2b = message.get('up2b', {}).get('nama_lokasi')
                else:
                    up2b = '' 

                if message.get('parent_lokasi'):
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi')
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = ''

                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 1, message['kode_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, parent_lokasi)
                ws_acc.write(index, 4, name_parent_lokasi)
                ws_acc.write(index, 5, message['lat'])
                ws_acc.write(index, 6, message['lon'])
                ws_acc.write(index, 7, up2b)
                ws_acc.write(index, 8, message['jenis_gi'])
                ws_acc.write(index, 9, message['fungsi_scada'])
                ws_acc.write(index, 10, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))