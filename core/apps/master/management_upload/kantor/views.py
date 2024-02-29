import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik, fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi
from apps.master.jaringan.ref_jenis_lokasi.models import RefJenisLokasi
from ..write_row import add_new_sheet
from .models import EXPORT_HEADERS, EXPORT_HEADERS_DATA 
class KantorView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi__in=[15,16,17,18,24,25])

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_kantor',
        summary="Management Upload SLD - Upload Temp Excel Kantor",
        description="Management Upload SLD - Upload Temp Excel Kantor",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_KANTOR']),
                    id_parent_lokasi=fillnan(value=row['ID_UNIT_INDUK']),
                    tree_jaringan=0,
                    id_ref_jenis_lokasi=fillnan(value=row['ID_JENIS_KANTOR']),
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Get Data Template Excel Kantor",
        description="Management Upload SLD - Get Data Template Excel Kantor",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='kantor')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_kantor.xlsx')
            workbook = xlsxwriter.Workbook(file_path) 
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            add_new_sheet(worksheet=data_upload,headers=EXPORT_HEADERS,format=format)   
 
            sample_upload.set_tab_color('red') 
            add_new_sheet(worksheet=sample_upload,headers=EXPORT_HEADERS,format=format)  
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_kantor'))
                sample_upload.write(index, 1, data.get('id_jenis_kantor'))
                sample_upload.write(index, 2, data.get('id_unit_induk'))
                sample_upload.write(index, 3, data.get('nama_kantor'))
                sample_upload.write(index, 4, data.get('alamat'))
                sample_upload.write(index, 5, data.get('latitude'))
                sample_upload.write(index, 6, data.get('longitude'))
                sample_upload.write(index, 7, data.get('status'))
                sample_upload.write(index, 8, data.get('event_upload')) 
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_UNIT_PEMBANGKIT UNTUK EVENT UPLOAD "INSERT"', merge_format)
            sample_upload.merge_range('A11:D11', 'WAJIB DIISIS ID_UNIT_INDUK UNTUK JENIS KANTOR (UP3, ULP, UP2D)  ', merge_format)

            ref_jenis_kantor = workbook.add_worksheet('REF JENIS_KANTOR') 
            ref_jenis_kantor.set_tab_color('red')
            ref_jenis_kantor.set_column('A:A', 15)
            ref_jenis_kantor.set_column('B:B', 25)
            ref_jenis_kantor.write('A1', "ID_JENIS_KANTOR", format)
            ref_jenis_kantor.write('B1', "NAMA_JENIS_KANTOR", format)

            jk = {
                'kantor':[
                     {
                        'id_jenis_kantor': 15, 
                        'nama_kantor':'UP3',
                     }, 
                     {
                        'id_jenis_kantor': 16, 
                        'nama_kantor':'ULP',
                     }, 
                     {
                        'id_jenis_kantor': 17, 
                        'nama_kantor':'UIW/UID',
                     }, 
                     {
                        'id_jenis_kantor': 18, 
                        'nama_kantor':'UP2D',
                     }, 
                     {
                        'id_jenis_kantor': 24, 
                        'nama_kantor':'UP3B',
                     }, 
                     {
                        'id_jenis_kantor': 25, 
                        'nama_kantor':'UP2B',
                     },  
                 ]
                }
            index = 1
            for message in jk.get('kantor'):
                ref_jenis_kantor.write(index, 0, str(message['id_jenis_kantor']))
                ref_jenis_kantor.write(index, 1, message['nama_kantor'])
                index += 1
            
            ref_uid = workbook.add_worksheet('REF UNIT INDUK')
            # ,24,25
            UID = RefLokasi.objects.filter(id_ref_jenis_lokasi=17).values()  
            if UID: 
                ref_uid.set_tab_color('red')
                ref_uid.set_column('A:A', 15)
                ref_uid.set_column('B:B', 25)
                ref_uid.write('A1', "ID_UNIT_INDUK", format)
                ref_uid.write('B1', "ID_JENIS_KANTOR", format)
                ref_uid.write('C1', "NAMA_JENIS_KANTOR", format)
                ref_uid.write('D1', "NAMA_KANTOR", format)
                index = 1
                for message in UID:   
                    jenis_lokasi = RefJenisLokasi.objects.values_list('nama_jenis_lokasi',flat=True).get(pk=int(message.get('id_ref_jenis_lokasi_id')))
                      
                    ref_uid.write(index, 0, str(message.get('id_ref_lokasi')))
                    ref_uid.write(index, 1, str(message.get('id_ref_jenis_lokasi_id')))
                    ref_uid.write(index, 2, jenis_lokasi)
                    ref_uid.write(index, 3, message.get('nama_lokasi'))
                    index += 1
 
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Get Data Excel Unit Pembangkit",
        description="Management Upload SLD - Get Data Excel Unit Pembangkit",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD UNIT PEMBANGKIT.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA UNIT PEMBANGKIT')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            add_new_sheet(worksheet=ws_acc,headers=EXPORT_HEADERS_DATA,format=format)   

            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('uid'):
                    uid = message.get('uid', {}).get('nama_lokasi')
                else:
                    uid = ''
                if message.get('up3_1'):
                    up3_1 = message.get('up3_1', {}).get('nama_lokasi')
                else:
                    up3_1 = ''
                if message.get('ref_jenis_lokasi'):
                    ref_jenis_lokasi = message.get('ref_jenis_lokasi', {}).get('nama_jenis_lokasi')
                else:
                    ref_jenis_lokasi = ''
                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, ref_jenis_lokasi)
                ws_acc.write(index, 4, message['id_parent_lokasi'])
                ws_acc.write(index, 5, uid)
                ws_acc.write(index, 6, up3_1)
                ws_acc.write(index, 7, message['lat'])
                ws_acc.write(index, 8, message['lon'])
                ws_acc.write(index, 9, status)
                # ws_acc.write(index, 5, uid)
                # ws_acc.write(index, 6, up3_1)
                # ws_acc.write(index, 7, ulp_1)
                # ws_acc.write(index, 8, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))