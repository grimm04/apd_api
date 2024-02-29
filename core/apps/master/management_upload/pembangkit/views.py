import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik,fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete

class PembangkitView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=2)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_pembangkit',
        summary="Management Upload SLD - Upload Temp Excel Pembangkit",
        description="Management Upload SLD - Upload Temp Excel Pembangkit",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_PEMBANGKIT']),
                    id_parent_lokasi=fillnan(value=row['ID_PARENT_LOKASI']),
                    id_ref_jenis_pembangkit=fillnan(value=row['ID_REF_JENIS_PEMBANGKIT']),
                    tree_jaringan=1,
                    id_ref_jenis_lokasi=2,
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    # id_uid=row['ID_UID'],
                    # id_up3_1=row['ID_UP3'],
                    # id_ulp_1=row['ID_ULP'],
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=row['EVENT_UPLOAD'],
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Pembangkit",
        description="Management Upload SLD - Download Template Excel Pembangkit",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='pembangkit')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_pembangkit.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            data_upload.set_column('A:A', 20)
            data_upload.set_column('B:B', 25)
            data_upload.set_column('C:C', 50)
            data_upload.set_column('D:D', 20)
            data_upload.set_column('E:E', 15)
            data_upload.set_column('F:F', 15)
            data_upload.set_column('G:G', 15)
            data_upload.set_column('H:H', 15)
            data_upload.set_column('I:I', 15)
            # data_upload.set_column('J:J', 15)
            # data_upload.set_column('K:K', 15)
            data_upload.write('A1', "ID_PEMBANGKIT", format)
            data_upload.write('B1', "NAMA_PEMBANGKIT", format)
            data_upload.write('C1', "ALAMAT", format)
            data_upload.write('D1', "ID_PARENT_LOKASI", format)
            data_upload.write('E1', "ID_REF_JENIS_PEMBANGKIT", format)
            data_upload.write('F1', "LATITUDE", format)
            data_upload.write('G1', "LONGITUDE", format)
            data_upload.write('H1', "STATUS", format) 
            data_upload.write('I1', "EVENT_UPLOAD", format) 
            # data_upload.write('G1', "ID_UID", format)
            # data_upload.write('H1', "ID_UP3", format)
            # data_upload.write('I1', "ID_ULP", format)
            # data_upload.write('J1', "STATUS", format)
            # data_upload.write('K1', "EVENT_UPLOAD", format)
            sample_upload.set_tab_color('red')
            sample_upload.set_column('A:A', 20)
            sample_upload.set_column('B:B', 25)
            sample_upload.set_column('C:C', 50)
            sample_upload.set_column('D:D', 20)
            sample_upload.set_column('E:E', 15)
            sample_upload.set_column('F:F', 20)
            sample_upload.set_column('G:G', 20)
            sample_upload.set_column('H:H', 15)
            sample_upload.set_column('I:I', 15)
            # sample_upload.set_column('J:J', 15)
            # sample_upload.set_column('K:K', 15)
            sample_upload.write('A1', "ID_PEMBANGKIT", format)
            sample_upload.write('B1', "NAMA_PEMBANGKIT", format)
            sample_upload.write('C1', "ALAMAT", format)
            sample_upload.write('D1', "ID_PARENT_LOKASI", format)
            sample_upload.write('E1', "ID_REF_JENIS_PEMBANGKIT", format)
            sample_upload.write('F1', "LATITUDE", format)
            sample_upload.write('G1', "LONGITUDE", format)
            sample_upload.write('H1', "STATUS", format)
            sample_upload.write('I1', "EVENT_UPLOAD", format)
            # sample_upload.write('G1', "ID_UID", format)
            # sample_upload.write('H1', "ID_UP3", format)
            # sample_upload.write('I1', "ID_ULP", format)
            # sample_upload.write('J1', "STATUS", format)
            # sample_upload.write('K1', "EVENT_UPLOAD", format)
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_pembangkit'))
                sample_upload.write(index, 1, data.get('nama_pembangkit'))
                sample_upload.write(index, 2, data.get('alamat'))
                sample_upload.write(index, 3, data.get('id_parent_lokasi'))
                sample_upload.write(index, 4, data.get('id_ref_jenis_pembangkit'))
                sample_upload.write(index, 5, data.get('latitude'))
                sample_upload.write(index, 6, data.get('longitude'))
                sample_upload.write(index, 7, data.get('status'))
                sample_upload.write(index, 8, data.get('event_upload'))
                # sample_upload.write(index, 6, data.get('id_uid'))
                # sample_upload.write(index, 7, data.get('id_up3_1'))
                # sample_upload.write(index, 8, data.get('id_ulp_1'))
                # sample_upload.write(index, 9, data.get('status'))
                # sample_upload.write(index, 10, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_PEMBANGKIT UNTUK EVENT UPLOAD "INSERT"', merge_format)

            _, message = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=1)
            if _: 
                ref_jenis_lokasi = workbook.add_worksheet('REF UNIT PEMBANGKIT')
                ref_jenis_lokasi.set_tab_color('red')
                ref_jenis_lokasi.set_column('A:A', 20)
                ref_jenis_lokasi.set_column('B:B', 25)
                ref_jenis_lokasi.write('A1', "ID_UNIT_PEMBANGKIT", format)
                ref_jenis_lokasi.write('B1', "NAMA_UNIT_PEMBANGKIT", format)
                index = 1
                for message in message:
                    ref_jenis_lokasi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_jenis_lokasi.write(index, 1, message['nama_lokasi'])
                    index += 1
                merge_format = workbook.add_format(
                    {
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'fg_color': 'red'
                    }
                )
                ref_jenis_lokasi.merge_range('D2:N2', 'ID_UNIT_PEMBANGKIT DIGUNAKAN UNTUK MENGISI ID_PARENT_LOKASI KETIKA UPLOAD', merge_format)
                
            ref_jenis_pembangkit = workbook.add_worksheet('REF JENIS PEMBANGKIT')
            _, jp = self.model.read_other_table(table='ref_jenis_pembangkit')
            if _: 
                ref_jenis_pembangkit.set_tab_color('red')
                ref_jenis_pembangkit.set_column('A:A', 15)
                ref_jenis_pembangkit.set_column('B:B', 25)
                ref_jenis_pembangkit.write('A1', "ID_REF_JENIS_PEMBANGKIT", format)
                ref_jenis_pembangkit.write('B1', "NAMA_REF_JENIS_PEMBANGKIT", format)
                index = 1
                for message in jp:
                    ref_jenis_pembangkit.write(index, 0, str(message['id_ref_jenis_pembangkit']))
                    ref_jenis_pembangkit.write(index, 1, message['nama'])
                    index += 1
                ref_jenis_pembangkit.merge_range('D2:N2', 'ID_REF_JENIS_PEMBANGKIT DIGUNAKAN UNTUK MENGISI ID_REF_JENIS_PEMBANGKIT KETIKA UPLOAD', merge_format)

            # ref_up2d = workbook.add_worksheet('REF UID')
            # _, UID = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=18)
            # if not _:
            #     return build_response(ApiResponse(message=str(UID)))
            # ref_up2d.set_tab_color('red')
            # ref_up2d.set_column('A:A', 15)
            # ref_up2d.set_column('B:B', 25)
            # ref_up2d.write('A1', "ID_UID", format)
            # ref_up2d.write('B1', "NAMA_UID", format)
            # index = 1
            # for message in UID:
            #     ref_up2d.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_up2d.write(index, 1, message['nama_lokasi'])
            #     index += 1

            # ref_up3 = workbook.add_worksheet('REF UP3')
            # _, UP3 = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=15)
            # if not _:
            #     return build_response(ApiResponse(message=str(UP3)))
            # ref_up3.set_tab_color('red')
            # ref_up3.set_column('A:A', 15)
            # ref_up3.set_column('B:B', 25)
            # ref_up3.write('A1', "ID_UP3", format)
            # ref_up3.write('B1', "NAMA_UP3", format)
            # index = 1
            # for message in UP3:
            #     ref_up3.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_up3.write(index, 1, message['nama_lokasi'])
            #     index += 1

            # ref_ulp = workbook.add_worksheet('REF ULP')
            # _, ULP = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=16)
            # if not _:
            #     return build_response(ApiResponse(message=str(ULP)))
            # ref_ulp.set_tab_color('red')
            # ref_ulp.set_column('A:A', 15)
            # ref_ulp.set_column('B:B', 25)
            # ref_ulp.write('A1', "ID_ULP", format)
            # ref_ulp.write('B1', "NAMA_ULP", format)
            # index = 1
            # for message in ULP:
            #     ref_ulp.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_ulp.write(index, 1, message['nama_lokasi'])
            #     index += 1

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Pembangkit",
        description="Management Upload SLD - Download Data Excel Pembangkit",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD MASTER PEMBANGKIT.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA PEMBANGKIT')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            ws_acc.set_column('A:A', 20)
            ws_acc.set_column('B:B', 25) 
            ws_acc.set_column('C:C', 50)
            ws_acc.set_column('D:D', 20)
            ws_acc.set_column('E:E', 25)
            ws_acc.set_column('F:F', 15)
            ws_acc.set_column('G:G', 15)
            ws_acc.set_column('H:H', 15)
            ws_acc.set_column('I:I', 15)
            ws_acc.set_column('J:J', 15)
            # ws_acc.set_column('K:K', 15)
            ws_acc.write('A1', "ID_PEMBANGKIT", format)
            ws_acc.write('B1', "NAMA_PEMBANGKIT", format)
            ws_acc.write('C1', "ALAMAT", format)
            ws_acc.write('D1', "ID_PARENT_LOKASI", format)
            ws_acc.write('E1', "NAMA_PARENT_LOKASI", format)
            ws_acc.write('F1', "ID_REF_JENIS_PEMBANGKIT", format)
            ws_acc.write('G1', "NAMA_UNIT_PEMBANGKIT", format)
            ws_acc.write('H1', "LATITUDE", format)
            ws_acc.write('I1', "LONGITUDE", format)
            ws_acc.write('J1', "STATUS", format)

            # ws_acc.write('H1', "UID", format)
            # ws_acc.write('I1', "UP3", format)
            # ws_acc.write('J1', "ULP", format)
            # ws_acc.write('K1', "STATUS", format)
            index = 1
            print(serializer.data)
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('ref_jenis_pembangkit'): 
                    jenis_pembangkit = message.get('ref_jenis_pembangkit', {}).get('nama')
                else:
                    jenis_pembangkit = ''
                # if message.get('up3_1'):
                #     up3_1 = message.get('up3_1', {}).get('nama_lokasi')
                # else:
                #     up3_1 = ''
                # if message.get('ulp_1'):
                #     ulp_1 = message.get('ulp_1', {}).get('nama_lokasi')
                # else:
                #     ulp_1 = '' 
                if message.get('parent_lokasi'):
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi')
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = ''

                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, parent_lokasi)
                ws_acc.write(index, 4, name_parent_lokasi)
                ws_acc.write(index, 5, message['id_ref_jenis_pembangkit'])
                ws_acc.write(index, 6, jenis_pembangkit)
                ws_acc.write(index, 7, message['lat'])
                ws_acc.write(index, 8, message['lon'])
                # ws_acc.write(index, 7, uid)
                # ws_acc.write(index, 8, up3_1)
                # ws_acc.write(index, 9, ulp_1)
                ws_acc.write(index, 9, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))