import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik, fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete
from ..write_row import add_new_sheet

from .models import EXPORT_HEADERS, EXPORT_HEADERS_DATA

class PenyulangView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=6)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_penyulang',
        summary="Management Upload SLD - Upload Temp Excel Penyulang",
        description="Management Upload SLD - Upload Temp Excel Penyulang",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            # print(data)
            data.fillna("nan", inplace=True)  
            for index, row in data.iterrows():
                id_gardu_induk  = None
                if row['ID_PARENT_LOKASI']:
                    rf = RefLokasi.objects.values_list('id_gardu_induk', flat=True).get(pk=int(row['ID_PARENT_LOKASI']))
                    id_gardu_induk =rf 
                data = dict(
                    kode_lokasi=fillnan(value=row['KODE_PENYULANG']),
                    nama_lokasi=fillnan(value=row['NAMA_PENYULANG']),
                    id_uid=fillnan(value=row['ID_UID']),
                    id_up3_1=fillnan(value=row['ID_UP3']),
                    id_ulp_1=fillnan(value=row['ID_ULP']),
                    id_parent_lokasi=fillnan(value=row['ID_PARENT_LOKASI']),
                    id_gardu_induk=id_gardu_induk,
                    tree_jaringan=1,
                    id_ref_jenis_lokasi=6,
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user), 
                    jenis_jaringan=fillnan(value=row['JENIS_JARINGAN']),
                    status_penyulang=fillnan(value=row['STATUS_PENYULANG']),
                    pemilik=fillnan(value=row['PEMILIK']), 
                    i_max=fillnan(value=row['I_MAX']), 
                    dcc=fillnan(value=row['DCC']), 
                    ratio_ct=fillnan(value=row['RATIO_CT']), 
                    ratio_vt=fillnan(value=row['RATIO_VT']), 
                    faktor_kali=fillnan(value=row['FAKTOR_KALI']), 
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    kva=fillnan(value=row['KVA']),
                    phase=fillnan(value=row['PHASE']), 
                    no_urut=fillnan(value=row['NO_URUT']), 
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Penyulang",
        description="Management Upload SLD - Download Template Excel Penyulang",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='penyulang')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_penyulang.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            add_new_sheet(worksheet=data_upload,headers=EXPORT_HEADERS,format=format) 
            data_upload.set_tab_color('green')

            add_new_sheet(worksheet=sample_upload,headers=EXPORT_HEADERS,format=format) 
            sample_upload.set_tab_color('red') 
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_penyulang'))
                sample_upload.write(index, 1, data.get('kode_lokasi'))
                sample_upload.write(index, 2, data.get('nama_penyulang'))
                sample_upload.write(index, 3, data.get('alamat'))
                sample_upload.write(index, 4, data.get('id_uid'))
                sample_upload.write(index, 5, data.get('id_up3_1'))
                sample_upload.write(index, 6, data.get('id_ulp'))
                sample_upload.write(index, 7, data.get('id_parent_lokasi'))
                sample_upload.write(index, 8, data.get('jenis_jaringan'))
                sample_upload.write(index, 9, data.get('status_penyulang'))
                sample_upload.write(index, 10, data.get('pemilik'))
                sample_upload.write(index, 11, data.get('i_max'))
                sample_upload.write(index, 12, data.get('dcc'))
                sample_upload.write(index, 13, data.get('ratio_ct'))
                sample_upload.write(index, 14, data.get('ratio_vt'))
                sample_upload.write(index, 15, data.get('faktor_kali'))
                sample_upload.write(index, 16, data.get('latitude'))
                sample_upload.write(index, 17, data.get('longitude'))
                sample_upload.write(index, 18, data.get('kva'))
                sample_upload.write(index, 19, data.get('phase'))
                sample_upload.write(index, 20, data.get('no_urut')) 
                sample_upload.write(index, 21, data.get('status'))
                sample_upload.write(index, 22, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_PENYULANG UNTUK EVENT UPLOAD "INSERT"', merge_format) 
            sample_upload.merge_range('A11:D11', 'HAPUS SHEET REFERENSI KETIKA AKAN MENG UPLOAD DATA', merge_format) 
            ref_uid = workbook.add_worksheet('REF UID')
            _, UID = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=17)
            if _: 
                ref_uid.set_tab_color('red')
                ref_uid.set_column('A:A', 15)
                ref_uid.set_column('B:B', 25)
                ref_uid.write('A1', "ID_UID", format)
                ref_uid.write('B1', "NAMA_UID", format)
                index = 1
                for message in UID:
                    ref_uid.write(index, 0, str(message['id_ref_lokasi']))
                    ref_uid.write(index, 1, message['nama_lokasi'])
                    index += 1

            ref_up3 = workbook.add_worksheet('REF UP3&UP2B')
            _, UP3 = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=15)
            if _: 
                ref_up3.set_tab_color('red')
                ref_up3.set_column('A:A', 15)
                ref_up3.set_column('B:B', 15)
                ref_up3.set_column('C:C', 25)
                ref_up3.write('A1', "ID_UP3", format)
                ref_up3.write('B1', "ID_UID", format)
                ref_up3.write('C1', "NAMA_UP3", format)
                index = 1
                for message in UP3:
                    ref_up3.write(index, 0, str(message['id_ref_lokasi']))
                    ref_up3.write(index, 1, str(message['id_uid']))
                    ref_up3.write(index, 2, message['nama_lokasi'])
                    index += 1

            ref_ulp = workbook.add_worksheet('REF ULP')
            _, ULP = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=16)
            if _: 
                ref_ulp.set_tab_color('red')
                ref_ulp.set_column('A:A', 15)
                ref_ulp.set_column('B:B', 25)
                ref_ulp.write('A1', "ID_ULP", format)
                ref_ulp.write('B1', "NAMA_ULP", format)
                index = 1
                for message in ULP:
                    ref_ulp.write(index, 0, str(message['id_ref_lokasi']))
                    ref_ulp.write(index, 1, message['nama_lokasi'])
                    index += 1


            ref_trafo_gi = workbook.add_worksheet('REF TRAFO GI')
            _, trafo_gi = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=5)
            if _: 
                ref_trafo_gi.set_tab_color('red')
                ref_trafo_gi.set_column('A:A', 15)
                ref_trafo_gi.set_column('B:B', 25)
                ref_trafo_gi.write('A1', "ID_TRAFO_GI", format)
                ref_trafo_gi.write('B1', "NAMA_TRAFO_GI", format)
                index = 1
                for message in trafo_gi:
                    ref_trafo_gi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_trafo_gi.write(index, 1, message['nama_lokasi'])
                    index += 1
                merge_format = workbook.add_format(
                    {
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'fg_color': 'red'
                    }
                )
                ref_trafo_gi.merge_range('D2:N2',
                                            'ID_TRAFO_GI DIGUNAKAN UNTUK MENGISI ID_PARENT LOKASI KETIKA UPLOAD',
                                            merge_format)

            #status_penyulang
            status_penyulang = workbook.add_worksheet('REF STATUS PENYULANG')  
            status_penyulang.set_tab_color('red')
            status_penyulang.set_column('A:A', 25) 
            status_penyulang.write('A1', "STATUS_PENYULANG", format)
            index = 1
            sp = ['BERBEBAN','EXPRESS','EXPRESS BERBEBAN','PEMBANGKIT','SPOTILOAD']
            for message in sp:
                status_penyulang.write(index, 0, str(message)) 
                index += 1
            
            #status_penyulang
            jenis_jaringan = workbook.add_worksheet('REF JENIS JARINGIN')  
            jenis_jaringan.set_tab_color('red')
            jenis_jaringan.set_column('A:A', 25) 
            jenis_jaringan.write('A1', "JENIS_JARINGIN", format)
            index = 1
            jj = ['CAMPURAN','SKTM','SUTM']
            for message in jj:
                jenis_jaringan.write(index, 0, message) 
                index += 1

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Penyulang",
        description="Management Upload SLD - Download Data Excel Penyulang",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'LAP MASTER PENYULANG.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA PENYULANG')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')

            add_new_sheet(worksheet=ws_acc,headers=EXPORT_HEADERS_DATA,format=format)  
            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('uid'):
                    uid = message.get('uid', {}).get('nama_lokasi')
                else:
                    uid = ''
                if message.get('up3_1'):
                    up3_1 = message.get('up3_1', {}).get('nama_lokasi')
                else:
                    up3_1 = ''
                if message.get('ulp_1'):
                    ulp_1 = message.get('ulp_1', {}).get('nama_lokasi')
                else:
                    ulp_1 = ''

                if message.get('parent_lokasi'):
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi')
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = '' 

                ws_acc.write(index, 0, message.get('id_ref_lokasi'))
                ws_acc.write(index, 1, message.get('kode_lokasi'))
                ws_acc.write(index, 2, message.get('nama_lokasi'))
                ws_acc.write(index, 3, message.get('alamat'))
                ws_acc.write(index, 4, uid)
                ws_acc.write(index, 5, up3_1)
                ws_acc.write(index, 6, ulp_1)
                ws_acc.write(index, 7, parent_lokasi)
                ws_acc.write(index, 8, name_parent_lokasi)
                ws_acc.write(index, 9, message.get('jenis_jaringan'))
                ws_acc.write(index, 10, message.get('status_penyulang'))
                ws_acc.write(index, 11, message.get('pemilik'))
                ws_acc.write(index, 12, message.get('i_max'))
                ws_acc.write(index, 13, message.get('dcc'))
                ws_acc.write(index, 14, message.get('ratio_ct'))
                ws_acc.write(index, 15, message.get('ratio_vt'))
                ws_acc.write(index, 16, message.get('faktor_kali'))
                ws_acc.write(index, 17, message.get('latitude'))
                ws_acc.write(index, 18, message.get('longitude'))
                ws_acc.write(index, 19, message.get('kva'))
                ws_acc.write(index, 20, message.get('phase'))
                ws_acc.write(index, 21, message.get('no_urut')) 
                ws_acc.write(index, 22, status) 
                index += 1

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))