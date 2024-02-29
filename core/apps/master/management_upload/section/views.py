import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik, fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete
from ..write_row import add_new_sheet
from .models import EXPORT_HEADERS, EXPORT_HEADERS_DATA 

class SectionView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=8)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_section',
        summary="Management Upload SLD - Upload Temp Excel Section",
        description="Management Upload SLD - Upload Temp Excel Section",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            } 
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                if row['ID_ZONE']:
                    rf = RefLokasi.objects.values_list('id_uid','id_up3_1','id_ulp_1').get(pk=int(row['ID_PENYULANG'])) 
                    id_uid  = rf[0]
                    id_up3_1  = rf[1]
                    id_ulp_1  = rf[2]
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_SECTION']),
                    id_gardu_induk=fillnan(value=row['ID_GARDU_INDUK']), 
                    id_trafo_gi=fillnan(value=row['ID_TRAFO_GI']),  
                    id_penyulang=fillnan(value=row['ID_PENYULANG']),
                    id_zone=fillnan(value=row['ID_ZONE']),
                    id_parent_lokasi=fillnan(value=row['ID_ZONE']),
                    tree_jaringan=1,
                    id_ref_jenis_lokasi=8,
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    id_uid  = id_uid,
                    id_up3_1  = id_up3_1,
                    id_ulp_1  = id_ulp_1,
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Section",
        description="Management Upload SLD - Download Template Excel Section",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='section')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_section.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            add_new_sheet(worksheet=data_upload,headers=EXPORT_HEADERS,format=format)   

            sample_upload.set_tab_color('red')
            add_new_sheet(worksheet=sample_upload,headers=EXPORT_HEADERS,format=format)   

            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_zona'))
                sample_upload.write(index, 1, data.get('nama_zona'))
                sample_upload.write(index, 2, data.get('alamat'))
                sample_upload.write(index, 3, data.get('id_gardu_induk'))
                sample_upload.write(index, 4, data.get('id_trafo_gi')) 
                sample_upload.write(index, 5, data.get('id_penyulang')) 
                sample_upload.write(index, 6, data.get('id_zone'))  
                sample_upload.write(index, 7, data.get('latitude'))
                sample_upload.write(index, 8, data.get('longitude')) 
                sample_upload.write(index, 9, data.get('status'))
                sample_upload.write(index, 10, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            ref_gi = workbook.add_worksheet('REF GARDU INDUK')
            _, GARDU_INDUK = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=4)
            if _: 
                ref_gi.set_tab_color('red')
                ref_gi.set_column('A:A', 15)
                ref_gi.set_column('B:B', 25)
                ref_gi.write('A1', "ID_GARDU_INDUK", format)
                ref_gi.write('B1', "NAMA_GARDU_INDUK", format)
                index = 1
                for message in GARDU_INDUK:
                    ref_gi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_gi.write(index, 1, message['nama_lokasi'])
                    index += 1
 
            ref_trafogi = workbook.add_worksheet('REF TRAFO GI')
            _, TRAFO_GI = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=5)
            if _: 
                ref_trafogi.set_tab_color('red')
                ref_trafogi.set_column('A:A', 15)
                ref_trafogi.set_column('B:B', 15)
                ref_trafogi.set_column('C:C', 25)
                ref_trafogi.write('A1', "ID_TRAFO_GI", format)
                ref_trafogi.write('B1', "ID_GARDU_INDUK", format)
                ref_trafogi.write('C1', "NAMA_TRAFO_GI", format)
                index = 1
                for message in TRAFO_GI:
                    ref_trafogi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_trafogi.write(index, 1, str(message['id_gardu_induk']))
                    ref_trafogi.write(index, 2, message['nama_lokasi'])
                    index += 1

            ref_penyulang = workbook.add_worksheet('REF PENYULANG')
            _, PENYULANG = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=6)
            if  _: 
                ref_penyulang.set_tab_color('red')
                ref_penyulang.set_column('A:A', 15)
                ref_penyulang.set_column('B:B', 25)
                ref_penyulang.set_column('C:C', 25)
                ref_penyulang.set_column('D:D', 25)
                ref_penyulang.write('A1', "ID_PENYULANG", format)
                ref_penyulang.write('B1', "ID_GARDU_INDUK", format)
                ref_penyulang.write('C1', "ID_TRAFO_GI", format)
                ref_penyulang.write('D1', "NAMA_PENYULANG", format)
                index = 1
                for message in PENYULANG:
                    ref_penyulang.write(index, 0, str(message['id_ref_lokasi']))
                    ref_penyulang.write(index, 1, str(message['id_gardu_induk']))
                    ref_penyulang.write(index, 2, str(message['id_parent_lokasi']))
                    ref_penyulang.write(index, 3, message['nama_lokasi'])
                    index += 1
                
            ref_zone = workbook.add_worksheet('REF ZONE')
            _, ZONE = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=7)
            if _: 
                ref_zone.set_tab_color('red')
                ref_zone.set_column('A:A', 15)
                ref_zone.set_column('B:B', 15)
                ref_zone.set_column('C:C', 25)
                ref_zone.set_column('D:D', 25)
                ref_zone.set_column('E:E', 25)
                ref_zone.write('A1', "ID_ZONE", format)
                ref_zone.write('B1', "ID_GARDU_INDUK", format)
                ref_zone.write('C1', "ID_TRAFO_GI", format)
                ref_zone.write('D1', "ID_PENYULANG", format)
                ref_zone.write('E1', "NAMA_PENYULANG", format)
                index = 1
                for message in ZONE:
                    ref_zone.write(index, 0, str(message['id_ref_lokasi']))
                    ref_zone.write(index, 1, str(message['id_gardu_induk']))
                    ref_zone.write(index, 2, str(message['id_trafo_gi']))
                    ref_zone.write(index, 3, str(message['id_penyulang']))
                    ref_zone.write(index, 4, message['nama_lokasi'])
                    index += 1
                merge_format = workbook.add_format(
                    {
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'fg_color': 'red'
                    }
                )
                ref_zone.merge_range('F2:O2', 'ID_ZONE DIGUNAKAN UNTUK MENGISI ID_ZONE KETIKA UPLOAD', merge_format)
                

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Section",
        description="Management Upload SLD - Download Data Excel Section",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD MASTER SECTION.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA SECTION')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            add_new_sheet(worksheet=ws_acc,headers=EXPORT_HEADERS_DATA,format=format)   

            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('parent_lokasi'): 
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi') 
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = '' 
                if message.get('trafo_gi'):
                    nama_trafo_gi= message.get('trafo_gi', {}).get('nama_lokasi') 
                else:
                    nama_trafo_gi= ''

                if message.get('gardu_induk'):
                    nama_gi= message.get('gardu_induk', {}).get('nama_lokasi') 
                else:
                    nama_gi= ''

                if message.get('trafo_gi'):
                    nama_trafo_gi= message.get('trafo_gi', {}).get('nama_lokasi') 
                else:
                    nama_trafo_gi= '' 
                if message.get('penyulang'):
                    nama_penyulang= message.get('penyulang', {}).get('nama_lokasi') 
                else:
                    nama_penyulang= '' 

                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, nama_gi)
                ws_acc.write(index, 4, nama_trafo_gi)
                ws_acc.write(index, 5, nama_penyulang)
                ws_acc.write(index, 6, parent_lokasi)
                ws_acc.write(index, 7, name_parent_lokasi)
                ws_acc.write(index, 8, message['lat'])
                ws_acc.write(index, 9, message['lon']) 
                ws_acc.write(index, 10, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))