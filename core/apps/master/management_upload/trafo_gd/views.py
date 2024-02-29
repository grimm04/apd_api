import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik, fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete

from ..write_row import add_new_sheet
from .models import EXPORT_HEADERS, EXPORT_HEADERS_DATA 
class TrafoGDView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=14)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_trafo_gd',
        summary="Management Upload SLD - Upload Temp Excel Trafo GD",
        description="Management Upload SLD - Upload Temp Excel Trafo GD",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                id_uid  =  None
                id_up3_1  =  None
                id_ulp_1  =  None
                if row['ID_GARDU_DISTRIBUSI']: 
                    rf = RefLokasi.objects.values_list('id_uid','id_up3_1','id_ulp_1').get(pk=int(row['ID_GARDU_DISTRIBUSI'])) 
                    id_uid  = rf[0]
                    id_up3_1  = rf[1]
                    id_ulp_1  = rf[2]
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_TRAFO_GD']),
                    id_gardu_induk=fillnan(value=row['ID_GARDU_INDUK']), 
                    id_trafo_gi=fillnan(value=row['ID_TRAFO_GI']),  
                    id_penyulang=fillnan(value=row['ID_PENYULANG']),
                    id_zone=fillnan(value=row['ID_ZONE']),
                    id_section=fillnan(value=row['ID_SECTION']),
                    id_segment=fillnan(value=row['ID_SEGMENT']),
                    id_gardu_distribusi=fillnan(value=row['ID_GARDU_DISTRIBUSI']),
                    id_parent_lokasi=fillnan(value=row['ID_GARDU_DISTRIBUSI']),
                    tree_jaringan=0,
                    id_ref_jenis_lokasi=14,
                    alamat=fillnan(value=row['ALAMAT']), 
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    id_uid  = id_uid,
                    id_up3_1  = id_up3_1,
                    id_ulp_1  = id_ulp_1,
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Trafo GD",
        description="Management Upload SLD - Download Template Excel Trafo GD",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='trafo_gd')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_trafo_gd.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            add_new_sheet(worksheet=data_upload,headers=EXPORT_HEADERS,format=format)   
 
            sample_upload.set_tab_color('red') 
            add_new_sheet(worksheet=sample_upload,headers=EXPORT_HEADERS,format=format)  
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_zona'))
                sample_upload.write(index, 1, data.get('nama_zona'))
                sample_upload.write(index, 2, data.get('alamat'))
                sample_upload.write(index, 3, data.get('id_gardu_induk'))
                sample_upload.write(index, 4, data.get('id_trafo_gi')) 
                sample_upload.write(index, 5, data.get('id_penyulang')) 
                sample_upload.write(index, 6, data.get('id_zone'))  
                sample_upload.write(index, 7, data.get('id_section'))  
                sample_upload.write(index, 8, data.get('id_segment'))  
                sample_upload.write(index, 9, data.get('id_gardu_distribusi'))  
                sample_upload.write(index, 10, data.get('latitude'))
                sample_upload.write(index, 11, data.get('longitude')) 
                sample_upload.write(index, 12, data.get('status'))
                sample_upload.write(index, 13, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_TRAFO_GD UNTUK EVENT UPLOAD "INSERT"', merge_format)
            sample_upload.merge_range('A11:D11', 'ID_GARDU_INDUK, ID_TRAFO_GI, ID_PENYULANG, ID_GARDU_DISTRIBUSI WAJIB DIISI', merge_format)

            ref_segment = workbook.add_worksheet('REF GARDU_DISTRIBUSI')
            _, GARDU_DISTRIBUSI = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=10)
            if _: 
                ref_segment.set_tab_color('red')
                ref_segment.set_column('A:A', 15)
                ref_segment.set_column('B:B', 15)
                ref_segment.set_column('C:C', 15)
                ref_segment.set_column('D:D', 15)
                ref_segment.set_column('E:E', 15)
                ref_segment.set_column('F:F', 15) 
                ref_segment.set_column('G:G', 15) 
                ref_segment.set_column('H:H', 25) 
                ref_segment.write('A1', "ID_GARDU_INDUK", format)
                ref_segment.write('B1', "ID_TRAFO_GI", format)
                ref_segment.write('C1', "ID_PENYULANG", format)
                ref_segment.write('D1', "ID_ZONE", format)
                ref_segment.write('E1', "ID_SECTION", format)
                ref_segment.write('F1', "ID_SEGMENT", format)
                ref_segment.write('G1', "ID_GARDU_DISTRIBUSI", format)
                ref_segment.write('H1', "NAMA_GARDU_DISTRIBUSI", format)
                index = 1
                for message in GARDU_DISTRIBUSI:
                    ref_segment.write(index, 0, str(message['id_gardu_induk']))
                    ref_segment.write(index, 1, str(message['id_trafo_gi']))
                    ref_segment.write(index, 2, str(message['id_penyulang']))
                    ref_segment.write(index, 3, str(message['id_zone']))  
                    ref_segment.write(index, 4, str(message['id_section']))  
                    ref_segment.write(index, 5, str(message['id_segment']))  
                    ref_segment.write(index, 6, str(message['id_ref_lokasi']))
                    ref_segment.write(index, 7, message['nama_lokasi'])
                    index += 1

            ref_segment = workbook.add_worksheet('REF SEGMENT')
            _, SEGMENT = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=9)
            if _: 
                ref_segment.set_tab_color('red')
                ref_segment.set_column('A:A', 15)
                ref_segment.set_column('B:B', 15)
                ref_segment.set_column('C:C', 15)
                ref_segment.set_column('D:D', 15)
                ref_segment.set_column('E:E', 15)
                ref_segment.set_column('F:F', 25) 
                ref_segment.write('A1', "ID_GARDU_INDUK", format)
                ref_segment.write('B1', "ID_TRAFO_GI", format)
                ref_segment.write('C1', "ID_PENYULANG", format)
                ref_segment.write('D1', "ID_ZONE", format)
                ref_segment.write('E1', "ID_SECTION", format)
                ref_segment.write('F1', "ID_SEGMENT", format)
                ref_segment.write('G1', "NAMA_SEGMENT", format)
                index = 1
                for message in SEGMENT:
                    ref_segment.write(index, 0, str(message['id_gardu_induk']))
                    ref_segment.write(index, 1, str(message['id_trafo_gi']))
                    ref_segment.write(index, 2, str(message['id_penyulang']))
                    ref_segment.write(index, 3, str(message['id_zone']))  
                    ref_segment.write(index, 4, str(message['id_section']))  
                    ref_segment.write(index, 5, str(message['id_ref_lokasi']))
                    ref_segment.write(index, 6, message['nama_lokasi'])
                    index += 1

            ref_gi = workbook.add_worksheet('REF GARDU INDUK')
            _, GARDU_INDUK = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=4)
            if _: 
                ref_gi.set_tab_color('red')
                ref_gi.set_column('A:A', 15)
                ref_gi.set_column('B:B', 25)
                ref_gi.write('A1', "ID_GARDU_INDUK", format)
                ref_gi.write('B1', "NAMA_GARDU_INDUK", format)
                index = 1
                for message in GARDU_INDUK:
                    ref_gi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_gi.write(index, 1, message['nama_lokasi'])
                    index += 1
 
            ref_trafogi = workbook.add_worksheet('REF TRAFO GI')
            _, TRAFO_GI = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=5)
            if _: 
                ref_trafogi.set_tab_color('red')
                ref_trafogi.set_column('A:A', 15)
                ref_trafogi.set_column('B:B', 15)
                ref_trafogi.set_column('C:C', 25)
                ref_trafogi.write('A1', "ID_TRAFO_GI", format)
                ref_trafogi.write('B1', "ID_GARDU_INDUK", format)
                ref_trafogi.write('C1', "NAMA_TRAFO_GI", format)
                index = 1
                for message in TRAFO_GI:
                    ref_trafogi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_trafogi.write(index, 1, str(message['id_gardu_induk']))
                    ref_trafogi.write(index, 2, message['nama_lokasi'])
                    index += 1

            ref_penyulang = workbook.add_worksheet('REF PENYULANG')
            _, PENYULANG = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=6)
            if  _: 
                ref_penyulang.set_tab_color('red')
                ref_penyulang.set_column('A:A', 15)
                ref_penyulang.set_column('B:B', 25)
                ref_penyulang.set_column('C:C', 25)
                ref_penyulang.set_column('D:D', 25)
                ref_penyulang.write('A1', "ID_PENYULANG", format)
                ref_penyulang.write('B1', "ID_GARDU_INDUK", format)
                ref_penyulang.write('C1', "ID_TRAFO_GI", format)
                ref_penyulang.write('D1', "NAMA_PENYULANG", format)
                index = 1
                for message in PENYULANG:
                    ref_penyulang.write(index, 0, str(message['id_ref_lokasi']))
                    ref_penyulang.write(index, 1, str(message['id_gardu_induk']))
                    ref_penyulang.write(index, 2, str(message['id_parent_lokasi']))
                    ref_penyulang.write(index, 3, message['nama_lokasi'])
                    index += 1 
            ref_zone = workbook.add_worksheet('REF ZONE')
            _, ZONE = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=7)
            if _: 
                ref_zone.set_tab_color('red')
                ref_zone.set_column('A:A', 15)
                ref_zone.set_column('B:B', 15)
                ref_zone.set_column('C:C', 25)
                ref_zone.set_column('D:D', 25)
                ref_zone.set_column('E:E', 25)
                ref_zone.write('A1', "ID_ZONE", format)
                ref_zone.write('B1', "ID_GARDU_INDUK", format)
                ref_zone.write('C1', "ID_TRAFO_GI", format)
                ref_zone.write('D1', "ID_PENYULANG", format)
                ref_zone.write('E1', "NAMA_PENYULANG", format)
                index = 1
                for message in ZONE:
                    ref_zone.write(index, 0, str(message['id_ref_lokasi']))
                    ref_zone.write(index, 1, str(message['id_gardu_induk']))
                    ref_zone.write(index, 2, str(message['id_trafo_gi']))
                    ref_zone.write(index, 3, str(message['id_penyulang']))
                    ref_zone.write(index, 4, message['nama_lokasi'])
                    index += 1
            
            ref_section = workbook.add_worksheet('REF SECTION')
            _, SECTION = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=8)
            if _: 
                ref_section.set_tab_color('red')
                ref_section.set_column('A:A', 15)
                ref_section.set_column('B:B', 15)
                ref_section.set_column('C:C', 15)
                ref_section.set_column('D:D', 15)
                ref_section.set_column('E:E', 15)
                ref_section.set_column('F:F', 25) 
                ref_section.write('A1', "ID_GARDU_INDUK", format)
                ref_section.write('B1', "ID_TRAFO_GI", format)
                ref_section.write('C1', "ID_PENYULANG", format)
                ref_section.write('D1', "ID_ZONE", format)
                ref_section.write('E1', "ID_SECTION", format)
                ref_section.write('F1', "NAMA_SECTION", format)
                index = 1
                for message in SECTION:
                    ref_section.write(index, 0, str(message['id_gardu_induk']))
                    ref_section.write(index, 1, str(message['id_trafo_gi']))
                    ref_section.write(index, 2, str(message['id_penyulang']))
                    ref_section.write(index, 3, str(message['id_zone']))  
                    ref_section.write(index, 4, str(message['id_ref_lokasi']))
                    ref_section.write(index, 5, message['nama_lokasi'])
                    index += 1

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Trafo GD",
        description="Management Upload SLD - Download Data Excel Trafo GD",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'LAP MASTER TRAFO GD.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA TRAFO GD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            ws_acc.set_column('A:A', 20)
            ws_acc.set_column('B:B', 25)
            ws_acc.set_column('B:B', 25)
            ws_acc.set_column('B:B', 25)
            ws_acc.set_column('C:C', 50)
            ws_acc.set_column('D:D', 20)
            ws_acc.set_column('E:E', 25)
            ws_acc.set_column('F:F', 15)
            ws_acc.set_column('G:G', 15)
            ws_acc.set_column('H:H', 50)
            ws_acc.set_column('I:I', 10)
            ws_acc.set_column('J:J', 10)
            ws_acc.set_column('K:K', 15)
            ws_acc.write('A1', "ID_TRAFO_GD", format)
            ws_acc.write('B1', "NAMA_TRAFO_GD", format)
            ws_acc.write('C1', "ALAMAT", format)
            ws_acc.write('D1', "ID_PARENT_LOKASI", format)
            ws_acc.write('E1', "NAMA_PARENT_LOKASI", format)
            ws_acc.write('F1', "LATITUDE", format)
            ws_acc.write('G1', "LONGITUDE", format)
            ws_acc.write('H1', "COVERAGE", format)
            ws_acc.write('I1', "KVA", format)
            ws_acc.write('J1', "PHASE", format)
            ws_acc.write('K1', "UID", format)
            ws_acc.write('L1', "UP3", format)
            ws_acc.write('M1', "ULP", format)
            ws_acc.write('N1', "STATUS", format)
            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                if message.get('uid'):
                    uid = message.get('uid', {}).get('nama_lokasi')
                else:
                    uid = ''
                if message.get('up3_1'):
                    up3_1 = message.get('up3_1', {}).get('nama_lokasi')
                else:
                    up3_1 = ''
                if message.get('ulp_1'):
                    ulp_1 = message.get('ulp_1', {}).get('nama_lokasi')
                else:
                    ulp_1 = ''

                if message.get('parent_lokasi'):
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi')
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = ''

                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, parent_lokasi)
                ws_acc.write(index, 4, name_parent_lokasi)
                ws_acc.write(index, 5, message['lat'])
                ws_acc.write(index, 6, message['lon'])
                ws_acc.write(index, 7, message['coverage'])
                ws_acc.write(index, 8, message['kva'])
                ws_acc.write(index, 9, message['phase'])
                ws_acc.write(index, 10, uid)
                ws_acc.write(index, 11, up3_1)
                ws_acc.write(index, 12, ulp_1)
                ws_acc.write(index, 13, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))