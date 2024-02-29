import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik, fillnan
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete

from ..write_row import add_new_sheet
from .models import EXPORT_HEADERS, EXPORT_HEADERS_DATA

class TrafoGIView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=5)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_trafo_gi',
        summary="Management Upload SLD - Upload Temp Excel Trafo GI",
        description="Management Upload SLD - Upload Temp Excel Trafo GI",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):\

        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str) 
            data = data.fillna('nan')   
            for index, row in data.iterrows(): 
                data = dict(
                    nama_lokasi=fillnan(value=row['NAMA_TRAFO_GI']),
                    id_parent_lokasi=fillnan(value=row['ID_PARENT_LOKASI']),
                    id_gardu_induk=fillnan(value=row['ID_PARENT_LOKASI']),
                    tree_jaringan=0,
                    id_ref_jenis_lokasi=5,
                    alamat=fillnan(value=row['ALAMAT']),
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    kapasitas=fillnan(value=row['KAPASITAS']),
                    sub_sistem=fillnan(value=row['SUB_SISTEM']),
                    pemilik=fillnan(value=row['PEMILIK']),
                    status_trafo=fillnan(value=row['STATUS_TRAFO']),
                    jenis_trafo=fillnan(value=row['JENIS_TRAFO']),
                    i_max=fillnan(value=row['I_MAX']),
                    ratio_ct=fillnan(value=row['RATIO_CT']),
                    ratio_vt=fillnan(value=row['RATIO_VT']),
                    fk_meter_pembanding=fillnan(value=row['FX_METER_PEMBANDING']),
                    def_pengukuran_teg_primer=fillnan(value=row['DEF_PENGUKURAN_TEG_PRIMER']),
                    def_pengukuran_teg_sekunder=fillnan(value=row['DEF_PENGUKURAN_TEG_SEKUNDER']),
                    def_nilai_cosq=fillnan(value=row['DEF_NILAI_COSQ']),
                    jenis_layanan=fillnan(value=row['JENIS_LAYANAN']), 
                    sinkron_data=fillnan(value=row['SINKRON_DATA']),
                    id_i=fillnan(value=row['ID_I']),
                    id_v=fillnan(value=row['ID_V']),
                    id_p=fillnan(value=row['ID_P']),
                    id_amr=fillnan(value=row['ID_AMR']) ,
                    id_portal_ext=fillnan(value=row['ID_PORTAL_EXT']),
                    url_webservice=fillnan(value=row['URL_WEBSERVICE']), 
                    lat=float(fillnan(value=row['LATITUDE'])) if fillnan(value=row['LATITUDE']) else None,
                    lon=float(fillnan(value=row['LONGITUDE'])) if fillnan(value=row['LONGITUDE']) else None,
                    coverage=fillnan(value=row['COVERAGE']),
                    kva=fillnan(value=row['KVA']),
                    phase=fillnan(value=row['PHASE']), 
                    status_listrik=status_listrik(fillnan(value=row['STATUS'])),
                    event_upload=fillnan(value=row['EVENT_UPLOAD']),
                )
                print(data)
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Template Excel Trafo GI",
        description="Management Upload SLD - Download Template Excel Trafo GI",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try: 
            data = template(name='trafo_gi')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_trafo_gi.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format() 
            format.set_bg_color('#D3D3D3')
            format.set_align('left')

            add_new_sheet(worksheet=data_upload,headers=EXPORT_HEADERS,format=format)
            data_upload.set_tab_color('green')  
            add_new_sheet(worksheet=sample_upload,headers=EXPORT_HEADERS,format=format)
            sample_upload.set_tab_color('red') 
            
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_trafo_gi'))
                sample_upload.write(index, 1, data.get('nama_trafo_gi'))
                sample_upload.write(index, 2, data.get('alamat'))
                sample_upload.write(index, 3, data.get('id_parent_lokasi'))
                sample_upload.write(index, 4, data.get('kapasitas'))
                sample_upload.write(index, 5, data.get('sub_sistem'))
                sample_upload.write(index, 6, data.get('pemilik'))
                sample_upload.write(index, 7, data.get('status_trafo'))
                sample_upload.write(index, 8, data.get('jenis_trafo'))
                sample_upload.write(index, 9, data.get('i_max'))
                sample_upload.write(index, 10, data.get('ratio_ct'))
                sample_upload.write(index, 11, data.get('ratio_vt'))
                sample_upload.write(index, 12, data.get('fk_meter_pembanding'))
                sample_upload.write(index, 13, data.get('def_pengukuran_teg_primer'))
                sample_upload.write(index, 14, data.get('def_pengukuran_teg_sekunder'))
                sample_upload.write(index, 15, data.get('def_nilai_cosq'))
                sample_upload.write(index, 16, data.get('jenis_layanan')) 
                sample_upload.write(index, 17, data.get('latitude'))
                sample_upload.write(index, 18, data.get('longitude'))
                sample_upload.write(index, 19, data.get('coverage'))
                sample_upload.write(index, 20, data.get('kva'))
                sample_upload.write(index, 21, data.get('phase')) 
                sample_upload.write(index, 22, data.get('status'))
                sample_upload.write(index, 23, data.get('sinkron_data'))
                sample_upload.write(index, 24, data.get('id_i'))
                sample_upload.write(index, 25, data.get('id_v'))
                sample_upload.write(index, 26, data.get('id_p'))
                sample_upload.write(index, 27, data.get('id_amr'))
                sample_upload.write(index, 28, data.get('id_portal_ext'))
                sample_upload.write(index, 29, data.get('url_webservice'))
                sample_upload.write(index, 30, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_TRAFO_GI UNTUK EVENT UPLOAD "INSERT"', merge_format)

            _, message = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=4)
            ref_jenis_lokasi = workbook.add_worksheet('REF GARDU INDUK')
            ref_jenis_lokasi.set_tab_color('red')
            ref_jenis_lokasi.set_column('A:A', 15)
            ref_jenis_lokasi.set_column('B:B', 25)
            ref_jenis_lokasi.write('A1', "ID_GARDU_INDUK", format)
            ref_jenis_lokasi.write('B1', "NAMA_GARDU_INDUK", format)
            if _: 
                index = 1
                for message in message:
                    ref_jenis_lokasi.write(index, 0, str(message['id_ref_lokasi']))
                    ref_jenis_lokasi.write(index, 1, message['nama_lokasi'])
                    index += 1
                merge_format = workbook.add_format(
                    {
                        'bold': 1,
                        'border': 1,
                        'align': 'center',
                        'valign': 'vcenter',
                        'fg_color': 'red'
                    }
                )
                ref_jenis_lokasi.merge_range('D2:N2', 'ID_GARDU_INDUK DIGUNAKAN UNTUK MENGISI ID_PARENT LOKASI KETIKA UPLOAD', merge_format)
            
 

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Download Data Excel Trafo GI",
        description="Management Upload SLD - Download Data Excel Trafo GI",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True) 
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD MASTER TRAFO GI.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA TRAFO GI')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            add_new_sheet(worksheet=ws_acc,headers=EXPORT_HEADERS_DATA,format=format)  
            index = 1
 
            for message in serializer.data: 
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active' 

                if message.get('parent_lokasi'):
                    parent_lokasi = message.get('parent_lokasi', {}).get('id_ref_lokasi')
                    name_parent_lokasi = message.get('parent_lokasi', {}).get('nama_lokasi')
                else:
                    parent_lokasi = ''
                    name_parent_lokasi = ''

                ws_acc.write(index, 0, message['id_ref_lokasi'])
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, parent_lokasi)
                ws_acc.write(index, 4, name_parent_lokasi)
                ws_acc.write(index, 5, message['kapasitas'])
                ws_acc.write(index, 6, message['sub_sistem'])
                ws_acc.write(index, 7, message['pemilik'])
                ws_acc.write(index, 8, message['status_trafo'])
                ws_acc.write(index, 9, message['jenis_trafo'])
                ws_acc.write(index, 10, message['i_max'])
                ws_acc.write(index, 11, message['ratio_ct'])
                ws_acc.write(index, 12, message['ratio_vt'])
                ws_acc.write(index, 13, message['fk_meter_pembanding'])
                ws_acc.write(index, 14, message['def_pengukuran_teg_primer'])
                ws_acc.write(index, 15, message['def_pengukuran_teg_sekunder'])
                ws_acc.write(index, 16, message['def_nilai_cosq'])
                ws_acc.write(index, 17, message['jenis_layanan']) 
                ws_acc.write(index, 18, message['lat'])
                ws_acc.write(index, 19, message['lon'])
                ws_acc.write(index, 20, message['coverage'])
                ws_acc.write(index, 21, message['kva'])
                ws_acc.write(index, 22, message['phase']) 
                ws_acc.write(index, 23, status)
                ws_acc.write(index, 24, message['sinkron_data'])
                ws_acc.write(index, 25, message['id_i'])
                ws_acc.write(index, 26, message['id_v'])
                ws_acc.write(index, 27, message['id_p'])
                ws_acc.write(index, 28, message['id_amr'])
                ws_acc.write(index, 29, message['id_portal_ext'])
                ws_acc.write(index, 30, message['url_webservice']) 
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))