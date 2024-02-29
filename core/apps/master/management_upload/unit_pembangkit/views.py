import os
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime
from django.http import HttpResponse, Http404
from django.conf import settings
from rest_framework import viewsets, renderers
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTTokenUserAuthentication
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter
from apps.master.jaringan.ref_lokasi import serializers
from apps.master.management_upload.models import managementUpload
from library.api_response import ApiResponse, build_response
from base.excel_template import template, status_listrik
from apps.master.jaringan.ref_lokasi.models import RefLokasi, RefLokasiTempDelete

class UnitPembangkitView(viewsets.GenericViewSet):

    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTTokenUserAuthentication]

    queryset = RefLokasi.objects.filter(id_ref_jenis_lokasi=1)

    serializer_class = serializers.RefLokasiSerializer
    create_serializer_class = serializers.CRRefLokasiSerializers
    update_serializer_class = serializers.UDRefLokasiSerializers

    model = managementUpload()

    @extend_schema(
        operation_id='upload_file_unit_pembangkit',
        summary="Management Upload SLD - Upload Temp Excel Unit Pembangkit",
        description="Management Upload SLD - Upload Temp Excel Unit Pembangkit",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary'
                    },
                    'id_user': {
                        'type': 'integer',
                        # 'format': 'integer'
                    }
                }
            }
        },
        tags=['master_management_upload']
    )
    @action(detail=False, methods=['post'], url_path='upload-excel-temporary', url_name='upload-excel-temporary')
    def upload_excel_temporary(self, request):
        try:
            file_obj = request.data['file']
            id_user = request.data['id_user']
            data = pd.read_excel(file_obj, dtype=str)
            data.fillna("nan", inplace=True)
            for index, row in data.iterrows():
                data = dict(
                    nama_lokasi=row['NAMA_UNIT_PEMBANGKIT'],
                    # id_parent_lokasi=0,
                    tree_jaringan=1,
                    id_ref_jenis_lokasi=1,
                    alamat=row['ALAMAT'],
                    id_user_entri=int(id_user),
                    id_user_update=int(id_user),
                    lat=float(row['LATITUDE']),
                    lon=float(row['LONGITUDE']),
                    # id_uid=row['ID_UID'],
                    # id_up3_1=row['ID_UP3'],
                    # id_ulp_1=row['ID_ULP'],
                    status_listrik=status_listrik(row['STATUS']),
                    event_upload=row['EVENT_UPLOAD'],
                )
                data['tgl_entri'] = datetime.now().strftime('%Y-%m-%d')
                _, message = self.model.create_data(data, 'ref_lokasi_temp_upload')
                if not _:
                    raise Exception('failed to save data: {}'.format(message))

            return build_response(ApiResponse(status=True, message='Success insert to table temprorary'))
        except Exception as e:
            return build_response(ApiResponse(message=str(e)))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Get Data Template Excel Unit Pembangkit",
        description="Management Upload SLD - Get Data Template Excel Unit Pembangkit",
    )
    @action(methods=['get'], detail=False, url_path='download-template', url_name='download-template')
    def download_template(self, request):
        try:
            data = template(name='unit_pembangkit')
            file_path = os.path.join(settings.MEDIA_ROOT, 'template_upload_master_unit_pembangkit.xlsx')
            workbook = xlsxwriter.Workbook(file_path) 
            data_upload = workbook.add_worksheet('DATA UPLOAD')
            sample_upload = workbook.add_worksheet('CONTOH UPLOAD')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            data_upload.set_tab_color('green')
            data_upload.set_column('A:A', 25)
            data_upload.set_column('B:B', 50)
            data_upload.set_column('C:C', 15)
            data_upload.set_column('D:D', 15)
            data_upload.set_column('E:E', 15)
            data_upload.set_column('F:F', 15)
            # data_upload.set_column('G:H', 15)
            # data_upload.set_column('H:H', 15)
            # data_upload.set_column('I:I', 15)
            # data_upload.set_column('J:J', 15)
            data_upload.write('A1', "ID_UNIT_PEMBANGKIT", format)
            data_upload.write('B1', "NAMA_UNIT_PEMBANGKIT", format)
            data_upload.write('C1', "ALAMAT", format)
            data_upload.write('D1', "LATITUDE", format)
            data_upload.write('E1', "LONGITUDE", format)
            data_upload.write('F1', "STATUS", format)
            data_upload.write('G1', "EVENT_UPLOAD", format)
            # data_upload.write('F1', "ID_UID", format)
            # data_upload.write('G1', "ID_UP3", format)
            # data_upload.write('H1', "ID_ULP", format)
            # data_upload.write('I1', "STATUS", format)
            # data_upload.write('J1', "EVENT_UPLOAD", format)
            sample_upload.set_tab_color('red')
            sample_upload.set_column('A:A', 25)
            sample_upload.set_column('B:B', 50)
            sample_upload.set_column('C:C', 15)
            sample_upload.set_column('D:D', 15)
            sample_upload.set_column('E:E', 15)
            sample_upload.set_column('F:F', 15)
            # sample_upload.set_column('G:G', 15)
            # sample_upload.set_column('H:H', 15)
            # sample_upload.set_column('I:I', 15)
            # sample_upload.set_column('J:J', 15)
            sample_upload.write('A1', "ID_UNIT_PEMBANGKIT", format)
            sample_upload.write('B1', "NAMA_UNIT_PEMBANGKIT", format)
            sample_upload.write('C1', "ALAMAT", format)
            sample_upload.write('D1', "LATITUDE", format)
            sample_upload.write('E1', "LONGITUDE", format)
            sample_upload.write('F1', "STATUS", format)
            sample_upload.write('G1', "EVENT_UPLOAD", format)
            # sample_upload.write('F1', "ID_UID", format)
            # sample_upload.write('G1', "ID_UP3", format)
            # sample_upload.write('H1', "ID_ULP", format)
            # sample_upload.write('I1', "STATUS", format)
            # sample_upload.write('J1', "EVENT_UPLOAD", format)
            example = data['examples']
            index = 1
            for data in example:
                sample_upload.write(index, 0, data.get('id_unit_pembangkit'))
                sample_upload.write(index, 1, data.get('nama_unit_pembangkit'))
                sample_upload.write(index, 2, data.get('alamat'))
                sample_upload.write(index, 3, data.get('latitude'))
                sample_upload.write(index, 4, data.get('longitude'))
                sample_upload.write(index, 5, data.get('status'))
                sample_upload.write(index, 6, data.get('event_upload'))
                # sample_upload.write(index, 5, data.get('id_uid'))
                # sample_upload.write(index, 6, data.get('id_up3_1'))
                # sample_upload.write(index, 7, data.get('id_ulp_1'))
                # sample_upload.write(index, 8, data.get('status'))
                # sample_upload.write(index, 9, data.get('event_upload'))
                index += 1
            merge_format = workbook.add_format(
                {
                    'bold': 1,
                    'border': 1,
                    'align': 'center',
                    'valign': 'vcenter',
                    'fg_color': 'red'
                }
            )
            sample_upload.merge_range('A10:D10', 'KOSONGKAN ID_UNIT_PEMBANGKIT UNTUK EVENT UPLOAD "INSERT"', merge_format)
            
            # ref_uid = workbook.add_worksheet('REF UID')
            # _, UID = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=18)
            # if not _:
            #     return build_response(ApiResponse(message=str(UID)))
            # ref_uid.set_tab_color('red')
            # ref_uid.set_column('A:A', 15)
            # ref_uid.set_column('B:B', 25)
            # ref_uid.write('A1', "ID_UID", format)
            # ref_uid.write('B1', "NAMA_UID", format)
            # index = 1
            # for message in UID:
            #     ref_uid.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_uid.write(index, 1, message['nama_lokasi'])
            #     index += 1

            # ref_up3 = workbook.add_worksheet('REF UP3')
            # _, UP3 = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=15)
            # if not _:
            #     return build_response(ApiResponse(message=str(UP3)))
            # ref_up3.set_tab_color('red')
            # ref_up3.set_column('A:A', 15)
            # ref_up3.set_column('B:B', 25)
            # ref_up3.write('A1', "ID_UP3", format)
            # ref_up3.write('B1', "NAMA_UP3", format)
            # index = 1
            # for message in UP3:
            #     ref_up3.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_up3.write(index, 1, message['nama_lokasi'])
            #     index += 1

            # ref_ulp = workbook.add_worksheet('REF ULP')
            # _, ULP = self.model.read_all_data(table='ref_lokasi', id_ref_jenis_lokasi=16)
            # if not _:
            #     return build_response(ApiResponse(message=str(ULP)))
            # ref_ulp.set_tab_color('red')
            # ref_ulp.set_column('A:A', 15)
            # ref_ulp.set_column('B:B', 25)
            # ref_ulp.write('A1', "ID_ULP", format)
            # ref_ulp.write('B1', "NAMA_ULP", format)
            # index = 1
            # for message in ULP:
            #     ref_ulp.write(index, 0, str(message['id_ref_lokasi']))
            #     ref_ulp.write(index, 1, message['nama_lokasi'])
            #     index += 1

            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))

    @extend_schema(
        tags=['master_management_upload'],
        summary="Management Upload SLD - Get Data Excel Unit Pembangkit",
        description="Management Upload SLD - Get Data Excel Unit Pembangkit",
    )
    @action(methods=['get'], detail=False, url_path='download-excel', url_name='download-excel')
    def download_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            if not serializer.data:
                return build_response(ApiResponse(message=str('empty')))
            file_path = os.path.join(settings.MEDIA_ROOT, 'SLD UNIT PEMBANGKIT.xlsx')
            workbook = xlsxwriter.Workbook(file_path)
            ws_acc = workbook.add_worksheet('DATA UNIT PEMBANGKIT')
            format = workbook.add_format()
            format.set_bg_color('#D3D3D3')
            format.set_align('left')
            ws_acc.set_column('A:A', 25)
            ws_acc.set_column('B:B', 25)
            ws_acc.set_column('C:C', 50)
            ws_acc.set_column('D:D', 15)
            ws_acc.set_column('E:E', 15)
            ws_acc.set_column('F:F', 15)
            # ws_acc.set_column('G:G', 15)
            # ws_acc.set_column('H:H', 15)
            # ws_acc.set_column('I:I', 15)
            # ws_acc.set_column('J:J', 15)
            ws_acc.write('A1', "ID_UNIT_PEMBANGKIT", format)
            ws_acc.write('B1', "NAMA_UNIT_PEMBANGKIT", format)
            ws_acc.write('C1', "ALAMAT", format)
            ws_acc.write('D1', "LATITUDE", format)
            ws_acc.write('E1', "LONGITUDE", format)
            ws_acc.write('F1', "STATUS", format)
            # ws_acc.write('F1', "UID", format)
            # ws_acc.write('G1', "UP3", format)
            # ws_acc.write('H1', "ULP", format)
            # ws_acc.write('I1', "STATUS", format)
            index = 1
            for message in serializer.data:
                status = message['status_listrik']
                if status == '0':
                    status = 'inactive'
                elif status == '1':
                    status = 'active'
                # if message.get('uid'):
                #     uid = message.get('uid', {}).get('nama_lokasi')
                # else:
                #     uid = ''
                # if message.get('up3_1'):
                #     up3_1 = message.get('up3_1', {}).get('nama_lokasi')
                # else:
                #     up3_1 = ''
                # if message.get('ulp_1'):
                #     ulp_1 = message.get('ulp_1', {}).get('nama_lokasi')
                # else:
                #     ulp_1 = ''
                ws_acc.write(index, 0, str(message['id_ref_lokasi']))
                ws_acc.write(index, 1, message['nama_lokasi'])
                ws_acc.write(index, 2, message['alamat'])
                ws_acc.write(index, 3, message['lat'])
                ws_acc.write(index, 4, message['lon'])
                ws_acc.write(index, 5, status)
                # ws_acc.write(index, 5, uid)
                # ws_acc.write(index, 6, up3_1)
                # ws_acc.write(index, 7, ulp_1)
                # ws_acc.write(index, 8, status)
                index += 1
            workbook.close()
            response = HttpResponse(open(file_path, 'rb'), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(file_path)
            return response
        except:
            return build_response(ApiResponse(message='files not found'))